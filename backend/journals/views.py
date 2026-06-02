"""Journal Entries CRUD + post/unpost + reports + bank-statement upload."""

import csv
import io
import re
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Sum, Q
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes, permission_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.models import Account, AccountType, SupplierAccountMapping
from .models import BankDescriptionMapping, BankStatementUpload, JournalEntry, JournalEntryLine


# ─── helpers ──────────────────────────────────────────────────────────────────

def _line_dict(line):
    return {
        "id": str(line.id),
        "account_id": str(line.account_id),
        "account_code": line.account.code,
        "account_name": line.account.name,
        "debit": str(line.debit),
        "credit": str(line.credit),
        "currency": line.currency,
        "exchange_rate": str(line.exchange_rate),
        "base_debit": str(line.base_debit),
        "base_credit": str(line.base_credit),
        "description": line.description,
    }


def _entry_dict(entry, include_lines=True):
    d = {
        "id": str(entry.id),
        "entry_number": entry.entry_number,
        "date": entry.date.isoformat() if hasattr(entry.date, 'isoformat') else str(entry.date),
        "description": entry.description,
        "reference": entry.reference,
        "source": entry.source,
        "source_id": entry.source_id,
        "is_posted": entry.is_posted,
        "total_debit": str(entry.total_debit),
        "total_credit": str(entry.total_credit),
        "is_balanced": entry.is_balanced,
        "created_by": entry.created_by,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
    }
    if include_lines:
        d["lines"] = [_line_dict(l) for l in entry.lines.select_related("account").all()]
    return d


def _next_entry_number():
    """Generate next sequential entry number like JE-0001."""
    last = JournalEntry.objects.order_by("-entry_number").first()
    if not last:
        return "JE-0001"
    try:
        num = int(last.entry_number.split("-")[1])
        return f"JE-{num + 1:04d}"
    except (IndexError, ValueError):
        return f"JE-{JournalEntry.objects.count() + 1:04d}"


# ─── Journal Entries CRUD ─────────────────────────────────────────────────────

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def journal_entries_list_create(request):
    if request.method == "GET":
        qs = JournalEntry.objects.all()
        # Filters
        if request.query_params.get("is_posted"):
            qs = qs.filter(is_posted=request.query_params["is_posted"].lower() == "true")
        if request.query_params.get("source"):
            qs = qs.filter(source=request.query_params["source"])
        if request.query_params.get("date_from"):
            qs = qs.filter(date__gte=request.query_params["date_from"])
        if request.query_params.get("date_to"):
            qs = qs.filter(date__lte=request.query_params["date_to"])
        if request.query_params.get("search"):
            q = request.query_params["search"]
            qs = qs.filter(
                Q(entry_number__icontains=q) |
                Q(description__icontains=q) |
                Q(reference__icontains=q)
            )

        total_count = qs.count()

        # Pagination
        try:
            page = int(request.query_params.get("page", 1))
        except (ValueError, TypeError):
            page = 1
        try:
            page_size = int(request.query_params.get("page_size", 50))
        except (ValueError, TypeError):
            page_size = 50
        page_size = min(page_size, 200)
        offset = (page - 1) * page_size
        entries = qs[offset:offset + page_size]

        return Response({
            "success": True,
            "data": {
                "journal_entries": [_entry_dict(e, include_lines=False) for e in entries],
                "count": total_count,
                "total_count": total_count,
                "page": page,
                "page_size": page_size,
            },
        })

    # POST — create journal entry with lines
    data = request.data
    if not data.get("date"):
        return Response(
            {"success": False, "message": "date is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    lines_data = data.get("lines", [])
    if len(lines_data) < 2:
        return Response(
            {"success": False, "message": "At least 2 lines required for double-entry"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Validate balance
    total_debit = Decimal("0")
    total_credit = Decimal("0")
    for line in lines_data:
        try:
            d = Decimal(str(line.get("debit", 0)))
            c = Decimal(str(line.get("credit", 0)))
        except (InvalidOperation, TypeError):
            return Response(
                {"success": False, "message": "Invalid debit/credit amount"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        total_debit += d
        total_credit += c

    if total_debit != total_credit:
        return Response(
            {"success": False, "message": f"Entry not balanced: DR {total_debit} != CR {total_credit}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Validate accounts exist
    account_ids = [line.get("account_id") for line in lines_data]
    accounts = {str(a.id): a for a in Account.objects.filter(id__in=account_ids)}
    for aid in account_ids:
        if aid not in accounts:
            return Response(
                {"success": False, "message": f"Account {aid} not found"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    with transaction.atomic():
        entry = JournalEntry.objects.create(
            entry_number=data.get("entry_number") or _next_entry_number(),
            date=data["date"],
            description=data.get("description", ""),
            reference=data.get("reference", ""),
            source=data.get("source", "manual"),
            source_id=data.get("source_id", ""),
            is_posted=False,
            created_by=f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
        )

        for line in lines_data:
            JournalEntryLine.objects.create(
                journal_entry=entry,
                account_id=line["account_id"],
                debit=Decimal(str(line.get("debit", 0))),
                credit=Decimal(str(line.get("credit", 0))),
                currency=line.get("currency", "MAD"),
                exchange_rate=Decimal(str(line.get("exchange_rate", 1))),
                description=line.get("description", ""),
            )

    entry.refresh_from_db()
    return Response({
        "success": True,
        "message": "Journal entry created",
        "data": {"journal_entry": _entry_dict(entry)},
    }, status=status.HTTP_201_CREATED)


@api_view(["GET", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def journal_entries_detail(request, pk):
    try:
        entry = JournalEntry.objects.get(id=pk)
    except JournalEntry.DoesNotExist:
        return Response(
            {"success": False, "message": "Journal entry not found"},
            status=status.HTTP_404_NOT_FOUND,
        )

    if request.method == "GET":
        return Response({
            "success": True,
            "data": {"journal_entry": _entry_dict(entry)},
        })

    if request.method == "DELETE":
        if entry.is_posted:
            return Response(
                {"success": False, "message": "Cannot delete a posted entry. Unpost it first."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        entry.delete()
        return Response({"success": True, "message": "Journal entry deleted"})

    # PUT — update
    if entry.is_posted:
        return Response(
            {"success": False, "message": "Cannot edit a posted entry. Unpost it first."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    data = request.data
    lines_data = data.get("lines", [])

    if lines_data:
        if len(lines_data) < 2:
            return Response(
                {"success": False, "message": "At least 2 lines required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        total_debit = sum(Decimal(str(l.get("debit", 0))) for l in lines_data)
        total_credit = sum(Decimal(str(l.get("credit", 0))) for l in lines_data)
        if total_debit != total_credit:
            return Response(
                {"success": False, "message": f"Entry not balanced: DR {total_debit} != CR {total_credit}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    with transaction.atomic():
        for field in ("date", "description", "reference", "source", "source_id"):
            if field in data:
                setattr(entry, field, data[field])
        if "entry_number" in data and data["entry_number"] != entry.entry_number:
            if JournalEntry.objects.filter(entry_number=data["entry_number"]).exclude(id=pk).exists():
                return Response(
                    {"success": False, "message": "Entry number already exists"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            entry.entry_number = data["entry_number"]
        entry.save()

        if lines_data:
            entry.lines.all().delete()
            for line in lines_data:
                JournalEntryLine.objects.create(
                    journal_entry=entry,
                    account_id=line["account_id"],
                    debit=Decimal(str(line.get("debit", 0))),
                    credit=Decimal(str(line.get("credit", 0))),
                    currency=line.get("currency", "MAD"),
                    exchange_rate=Decimal(str(line.get("exchange_rate", 1))),
                    description=line.get("description", ""),
                )

    return Response({
        "success": True,
        "message": "Journal entry updated",
        "data": {"journal_entry": _entry_dict(entry)},
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def journal_entry_post(request, pk):
    try:
        entry = JournalEntry.objects.get(id=pk)
    except JournalEntry.DoesNotExist:
        return Response(
            {"success": False, "message": "Journal entry not found"},
            status=status.HTTP_404_NOT_FOUND,
        )

    if entry.is_posted:
        return Response(
            {"success": False, "message": "Already posted"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not entry.is_balanced:
        return Response(
            {"success": False, "message": "Cannot post unbalanced entry"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if entry.lines.count() < 2:
        return Response(
            {"success": False, "message": "Entry must have at least 2 lines"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    entry.is_posted = True
    entry.save(update_fields=["is_posted", "updated_at"])
    return Response({
        "success": True,
        "message": "Journal entry posted",
        "data": {"journal_entry": _entry_dict(entry)},
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def journal_entry_unpost(request, pk):
    try:
        entry = JournalEntry.objects.get(id=pk)
    except JournalEntry.DoesNotExist:
        return Response(
            {"success": False, "message": "Journal entry not found"},
            status=status.HTTP_404_NOT_FOUND,
        )

    if not entry.is_posted:
        return Response(
            {"success": False, "message": "Entry is not posted"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    entry.is_posted = False
    entry.save(update_fields=["is_posted", "updated_at"])
    return Response({
        "success": True,
        "message": "Journal entry unposted",
        "data": {"journal_entry": _entry_dict(entry)},
    })


# ─── Bank Statement Upload ────────────────────────────────────────────────────

def _parse_date(val):
    """Parse a date value from CSV/Excel into a date object."""
    if hasattr(val, "date"):
        return val.date()
    if hasattr(val, "strftime"):
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _safe_decimal(val):
    """Convert a value to Decimal, returning 0 for empty/invalid.

    Handles European formats like 1.500,50 and 1 500,50 as well as
    standard formats like 1,500.50 and 1500.50.
    """
    if val is None:
        return Decimal("0")
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    val = str(val).strip().replace("\xa0", "").replace(" ", "")
    if not val:
        return Decimal("0")
    # Detect European format: comma is decimal separator
    # e.g. "1.500,50" or "1500,50"
    if "," in val and "." in val:
        if val.rindex(",") > val.rindex("."):
            # European: 1.500,50 → 1500.50
            val = val.replace(".", "").replace(",", ".")
        else:
            # US: 1,500.50 → 1500.50
            val = val.replace(",", "")
    elif "," in val:
        # Could be "1500,50" (European decimal) or "1,500" (US thousands)
        # If exactly 3 digits after comma, treat as thousands separator
        parts = val.split(",")
        if len(parts) == 2 and len(parts[1]) == 3 and parts[1].isdigit():
            val = val.replace(",", "")
        else:
            val = val.replace(",", ".")
    try:
        return Decimal(val)
    except InvalidOperation:
        return Decimal("0")


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def bank_statement_upload(request):
    """Parse an uploaded bank statement file and return transactions as JSON."""
    uploaded = request.FILES.get("file")
    if not uploaded:
        return Response(
            {"success": False, "message": "No file uploaded"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    bank_account_id = request.data.get("bank_account_id")
    if not bank_account_id:
        return Response(
            {"success": False, "message": "bank_account_id is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Verify bank account exists
    try:
        Account.objects.get(id=bank_account_id)
    except Account.DoesNotExist:
        return Response(
            {"success": False, "message": "Bank account not found"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    name = uploaded.name.lower()
    transactions = []
    file_format = "raw"

    # Pre-load all accounts by code for fast lookup (no per-row queries)
    account_by_code = {a.code: a for a in Account.objects.filter(is_active=True)}

    try:
        if name.endswith(".csv"):
            content = uploaded.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                # Map CSV columns (case-insensitive lookup)
                cols = {k.strip().lower(): v for k, v in row.items()}
                op_date = cols.get("operation date") or cols.get("date") or cols.get("dt opération") or cols.get("dt operation")
                description = cols.get("detailed description") or cols.get("description") or cols.get("libellé large") or cols.get("libelle large") or cols.get("libellé") or ""
                reference = cols.get("reference") or cols.get("référence") or ""
                debit = _safe_decimal(cols.get("debit") or cols.get("débit"))
                credit = _safe_decimal(cols.get("credit") or cols.get("crédit"))

                parsed_date = _parse_date(op_date)
                if not parsed_date:
                    continue
                if debit == 0 and credit == 0:
                    continue

                txn = {
                    "date": parsed_date.isoformat(),
                    "description": str(description).strip(),
                    "reference": str(reference).strip(),
                    "debit": str(debit),
                    "credit": str(credit),
                    "account_id": None,
                }
                # Check for account_code column (pre-categorized)
                acct_code = cols.get("account_code") or cols.get("account code")
                if acct_code:
                    file_format = "categorized"
                    acct_code = str(acct_code).strip()
                    acct = account_by_code.get(acct_code)
                    if acct:
                        txn["account_id"] = str(acct.id)
                        txn["account_code"] = acct.code
                        txn["account_name"] = acct.name
                    else:
                        txn["account_code"] = acct_code
                        txn["account_name"] = f"(unknown: {acct_code})"

                transactions.append(txn)

        elif name.endswith(".xlsx") or name.endswith(".xls"):
            # Read headers and data rows from either format
            if name.endswith(".xls"):
                import xlrd
                raw_bytes = uploaded.read()
                xls_wb = xlrd.open_workbook(file_contents=raw_bytes)
                xls_ws = xls_wb.sheet_by_index(0)
                headers = [str(xls_ws.cell_value(0, c) or "").strip().lower() for c in range(xls_ws.ncols)]
                data_rows = []
                for r in range(1, xls_ws.nrows):
                    row_vals = []
                    for c in range(xls_ws.ncols):
                        cell = xls_ws.cell(r, c)
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            dt_tuple = xlrd.xldate_as_tuple(cell.value, xls_wb.datemode)
                            from datetime import date as _date
                            row_vals.append(_date(dt_tuple[0], dt_tuple[1], dt_tuple[2]))
                        else:
                            row_vals.append(cell.value)
                    data_rows.append(row_vals)
            else:
                import openpyxl
                wb = openpyxl.load_workbook(uploaded, data_only=True)
                ws = wb.active
                headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
                data_rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]

            # Detect if this is a categorized file (has account_code column)
            has_acct = "account_code" in headers or "account code" in headers

            # Map column indices
            col_map = {}
            for i, h in enumerate(headers):
                if h in ("operation date", "date", "dt opération", "dt operation"):
                    col_map["date"] = i
                elif h in ("detailed description", "description", "libellé large", "libelle large", "libellé"):
                    col_map["description"] = i
                elif h in ("reference", "référence"):
                    col_map["reference"] = i
                elif h in ("debit", "débit"):
                    col_map["debit"] = i
                elif h in ("credit", "crédit"):
                    col_map["credit"] = i
                elif h in ("account_code", "account code"):
                    col_map["account_code"] = i

            if "date" not in col_map:
                return Response(
                    {"success": False, "message": "Could not find date column in Excel file. Expected 'Operation Date' or 'Date'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for vals in data_rows:
                op_date = vals[col_map["date"]] if col_map.get("date") is not None else None
                description = vals[col_map["description"]] if col_map.get("description") is not None else ""
                reference = vals[col_map["reference"]] if col_map.get("reference") is not None else ""
                debit = _safe_decimal(vals[col_map["debit"]] if col_map.get("debit") is not None else None)
                credit = _safe_decimal(vals[col_map["credit"]] if col_map.get("credit") is not None else None)

                parsed_date = _parse_date(op_date)
                if not parsed_date:
                    continue
                if debit == 0 and credit == 0:
                    continue

                txn = {
                    "date": parsed_date.isoformat(),
                    "description": str(description or "").strip(),
                    "reference": str(reference or "").strip(),
                    "debit": str(debit),
                    "credit": str(credit),
                    "account_id": None,
                }

                if has_acct and col_map.get("account_code") is not None:
                    file_format = "categorized"
                    acct_code = str(vals[col_map["account_code"]] or "").strip()
                    if acct_code:
                        acct = account_by_code.get(acct_code)
                        if acct:
                            txn["account_id"] = str(acct.id)
                            txn["account_code"] = acct.code
                            txn["account_name"] = acct.name
                        else:
                            txn["account_code"] = acct_code
                            txn["account_name"] = f"(unknown: {acct_code})"

                transactions.append(txn)
        else:
            return Response(
                {"success": False, "message": "Unsupported file format. Use .csv or .xlsx"},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except Exception as e:
        return Response(
            {"success": False, "message": f"Error parsing file: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Duplicate detection: check for existing bank-source entries in the same date range
    duplicate_warning = ""
    if transactions:
        dates = [t["date"] for t in transactions]
        min_date = min(dates)
        max_date = max(dates)
        existing_count = JournalEntry.objects.filter(
            source="bank",
            date__gte=min_date,
            date__lte=max_date,
        ).count()
        if existing_count > 0:
            duplicate_warning = (
                f"Warning: {existing_count} bank-source journal entries already exist "
                f"between {min_date} and {max_date}. This upload may create duplicates."
            )

    return Response({
        "success": True,
        "data": {
            "transactions": transactions,
            "format": file_format,
            "duplicate_warning": duplicate_warning,
        },
    })


# French bank description patterns → used to extract entity names
_ENTITY_PATTERNS = [
    # "VIREMENT EMIS EN FAVEUR DE HOTEL+PERLE+DU+SUD" → "hotel perle du sud"
    re.compile(r"(?:virement|vir\.?\s*\w*)\s+(?:\w+\s+)*en\s+faveur\s+de\s+(.+)", re.I),
    # "VIREMENT RECU DE V M MOROCCO" → "v m morocco"
    re.compile(r"virement\s+recu\s+de\s+(.+)", re.I),
    # "RECEPTION RAPATRIEMENT EN FAVEUR DE Viktors Farmor A/S" → "viktors farmor a/s"
    re.compile(r"reception\s+rapatriement\s+en\s+faveur\s+de\s+(.+)", re.I),
    # "ACHAT PAR CARTE DE PAIEMENT CHEZ JARDIN MAJORELLE" → "jardin majorelle"
    re.compile(r"achat\s+par\s+carte.*?chez\s+(.+)", re.I),
    # "PAIEMENT INTERNET NATIONAL CARTE 7525 RAM INTERNET" → "ram internet"
    re.compile(r"paiement\s+internet\s+national\s+carte\s+\w+\s+(.+)", re.I),
    # "PAIEMENT PRLV MAROC TELECOM 28022026" → "maroc telecom"
    re.compile(r"paiement\s+prlv\s+(.+?)(?:\s+\d{6,}|$)", re.I),
]

# French bank keywords → fixed account codes
_KEYWORD_ACCOUNT_MAP = {
    # Bank fees / commissions
    "commission virement": "598001",       # Transfer Fee
    "commission rapatriement": "598001",   # Transfer Fee
    "commission prelevement": "598003",    # Bank Charges
    "commission tenue": "598002",          # Bank Account Fee
    "commission": "598003",               # Bank Charges (generic)
    "frais bancaire": "598003",
    "frais de tenue": "598002",
    "frais prelevement": "598002",        # Bank Account Fee
    "agios": "598000",                    # Interest Expenses
    "interet debiteur": "598000",
}

# Entity name aliases → account code (for names that don't match directly)
_ENTITY_ALIAS_MAP = {
    "ram internet": "533500",              # Royal Air Maroc → Flight Ticket
    "ram": "533500",
    "royal air maroc": "533500",
    "oncf": "553819",                      # Train
    "daoud naoufal": "513000",             # La Caisse Entrees
    "comptable": "592307",                 # Comptabilite
}


def _normalize_entity(raw):
    """Normalize an entity name extracted from bank description."""
    name = raw.strip()
    # Remove trailing reference numbers, dates
    name = re.sub(r"\s+REF\s+\w*\d+.*$", "", name, flags=re.I)
    name = re.sub(r"\s+\d{6,}.*$", "", name)
    # Replace + with space (bank format: HOTEL+PERLE+DU+SUD)
    name = name.replace("+", " ")
    # Remove trailing E-, E-CO etc (truncated merchant names)
    name = re.sub(r"\s+E-?\w*$", "", name)
    return name.strip().lower()


def _extract_description_key(description):
    """Extract a stable key from a bank description for learning.

    Uses the entity patterns to extract the entity name, which stays the same
    across transactions even when reference numbers change.
    Falls back to the full normalized description.
    """
    desc = (description or "").strip()
    if not desc:
        return ""
    for pattern in _ENTITY_PATTERNS:
        m = pattern.search(desc)
        if m:
            return _normalize_entity(m.group(1))
    # Fallback: remove numbers/references and normalize
    key = desc.lower()
    key = re.sub(r"\s+\d{6,}", "", key)
    key = key.replace("+", " ")
    key = re.sub(r"\s+", " ", key).strip()
    return key


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def bank_statement_auto_map(request):
    """Suggest account mappings for bank statement transactions.

    Understands French bank statement patterns:
    - COMMISSION ... → bank fee accounts (598xxx)
    - VIREMENT EMIS EN FAVEUR DE <name> → supplier (match against account names)
    - RECEPTION RAPATRIEMENT EN FAVEUR DE <name> → client (41xxxx)
    - PAIEMENT INTERNET NATIONAL CARTE ... <merchant> → supplier
    - ACHAT PAR CARTE DE PAIEMENT CHEZ <merchant> → supplier
    - PAIEMENT PRLV <company> → operating expense
    - VIREMENT COMMERCIAL RECU → client (generic)
    """
    txns = request.data.get("transactions", [])
    bank_account_id = request.data.get("bank_account_id")
    if not txns:
        return Response({"success": True, "data": {"suggestions": []}})

    SYSTEM_CODES = {"240000", "210000", "420000"}

    # Load learned mappings (highest priority)
    learned = {}  # description_key → {account_id, code, name}
    for m in BankDescriptionMapping.objects.select_related("account").all():
        learned[m.description_key] = {
            "account_id": str(m.account_id),
            "account_code": m.account.code,
            "account_name": m.account.name,
        }

    # Load all non-bank, non-system accounts
    all_accounts = list(
        Account.objects.filter(is_active=True)
        .exclude(code__startswith="10")
        .exclude(code__in=SYSTEM_CODES)
        .values("id", "code", "name")
    )
    acct_by_code = {}
    for a in all_accounts:
        acct_by_code[a["code"]] = a

    # Find parent accounts (accounts that have children) — prefer children over parents
    parent_ids = set(
        str(v) for v in
        Account.objects.filter(parent__isnull=False)
        .values_list("parent_id", flat=True).distinct()
    )

    # Build name lookup: normalized name → account dict
    # "hotel perle du sud" → {id, code, name}
    name_index = {}  # normalized_name → account
    word_index = {}  # word → list of (account, total_words_in_name)
    for a in all_accounts:
        norm = a["name"].lower().strip()
        name_index[norm] = a
        words = norm.split()
        is_parent = str(a["id"]) in parent_ids
        for w in words:
            if len(w) > 2:
                word_index.setdefault(w, []).append((a, len(words), is_parent))

    def _make_suggestion(a):
        return {"account_id": str(a["id"]), "account_code": a["code"], "account_name": a["name"]}

    def _match_entity(entity_name):
        """Try to match an extracted entity name against account names."""
        if not entity_name:
            return None
        norm = entity_name.lower().strip()

        # 0) Check alias map first
        if norm in _ENTITY_ALIAS_MAP:
            acct = acct_by_code.get(_ENTITY_ALIAS_MAP[norm])
            if acct:
                return acct

        # 1) Exact name match
        if norm in name_index:
            return name_index[norm]

        # 2) Check if entity is contained in an account name or vice versa
        #    Skip parent accounts — prefer specific children
        for acct_name, acct in name_index.items():
            if str(acct["id"]) in parent_ids:
                continue
            if norm in acct_name or acct_name in norm:
                return acct

        # 3) Word scoring with prefix matching
        #    "hachimi moha" should match "Guide hachimi mohamed" (moha→mohamed)
        entity_words = [w for w in norm.split() if len(w) > 2]
        if not entity_words:
            return None
        scores = {}  # account_id → [match_count, total_words, acct, is_parent]
        for w in entity_words:
            # Exact word match
            if w in word_index:
                for acct, total_words, is_parent in word_index[w]:
                    aid = str(acct["id"])
                    if aid not in scores:
                        scores[aid] = [0, total_words, acct, is_parent]
                    scores[aid][0] += 1
            # Prefix match: "moha" matches "mohamed", "seban" matches "sebban"
            for kw, entries in word_index.items():
                if kw == w:
                    continue
                if kw.startswith(w) or w.startswith(kw):
                    for acct, total_words, is_parent in entries:
                        aid = str(acct["id"])
                        if aid not in scores:
                            scores[aid] = [0, total_words, acct, is_parent]
                        scores[aid][0] += 0.5

        if not scores:
            return None
        # Best = most matches, prefer non-parent (leaf), tiebreak fewer words
        best_id = max(
            scores,
            key=lambda k: (scores[k][0], not scores[k][3], -scores[k][1]),
        )
        match_count = scores[best_id][0]
        if match_count >= 1:
            return scores[best_id][2]
        return None

    # ── Map each transaction ──
    suggestions = []
    for i, txn in enumerate(txns):
        desc = (txn.get("description") or "").strip()
        desc_lower = desc.lower()
        if not desc_lower:
            suggestions.append(None)
            continue

        matched = None

        # Step 0: Check learned mappings (highest priority)
        desc_key = _extract_description_key(desc)
        if desc_key and desc_key in learned:
            lm = learned[desc_key]
            suggestions.append({
                "index": i,
                "account_id": lm["account_id"],
                "account_code": lm["account_code"],
                "account_name": lm["account_name"],
            })
            continue

        # Step 1: Check fixed keyword patterns (commissions, fees)
        for keyword, code in _KEYWORD_ACCOUNT_MAP.items():
            if keyword in desc_lower:
                acct = acct_by_code.get(code)
                if acct:
                    matched = acct
                break

        # Step 2: Extract entity name from French bank patterns and match
        if not matched:
            for pattern in _ENTITY_PATTERNS:
                m = pattern.search(desc)
                if m:
                    entity = _normalize_entity(m.group(1))
                    matched = _match_entity(entity)
                    break

        # Step 3: If still no match, try matching the full description against account names
        if not matched:
            matched = _match_entity(desc_lower)

        if matched:
            suggestions.append({
                "index": i,
                "account_id": str(matched["id"]),
                "account_code": matched["code"],
                "account_name": matched["name"],
            })
        else:
            suggestions.append(None)

    return Response({
        "success": True,
        "data": {"suggestions": suggestions},
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def bank_statement_confirm(request):
    """Create journal entries from confirmed bank statement transactions.

    Uses the same margin-scheme classification as import_tva_q1:
    - Credit + 41xxxx (client)     → Pattern A: DR Bank, CR Client Funds (240000)
    - Credit + other               → DR Bank, CR selected account
    - Debit + COS descendant       → Pattern B: DR Client Funds (240000), CR Bank
    - Debit + 41xxxx (client refund) → Pattern B: DR Client Funds (240000), CR Bank
    - Debit + 10xxxx (bank/cash)   → Asset transfer: DR selected account, CR Bank
    - Debit + other (opex)         → Pattern C: DR selected account, CR Bank
    """
    data = request.data
    bank_account_id = data.get("bank_account_id")
    txns = data.get("transactions", [])

    if not bank_account_id:
        return Response(
            {"success": False, "message": "bank_account_id is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if not txns:
        return Response(
            {"success": False, "message": "No transactions to create"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        bank_account = Account.objects.get(id=bank_account_id)
    except Account.DoesNotExist:
        return Response(
            {"success": False, "message": "Bank account not found"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Split into mapped (have account) and all (for upload history)
    all_txns = txns
    mapped_txns = [t for t in txns if t.get("account_id")]

    account_ids = [t.get("account_id") for t in mapped_txns]
    accounts_map = {}
    if account_ids:
        accounts_map = {str(a.id): a for a in Account.objects.filter(id__in=account_ids)}
        for aid in account_ids:
            if aid not in accounts_map:
                return Response(
                    {"success": False, "message": f"Account {aid} not found"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

    # Client Funds Liability account — required for margin scheme
    try:
        client_funds_acct = Account.objects.get(code="240000")
    except Account.DoesNotExist:
        return Response(
            {"success": False, "message": "Account 240000 (Client Funds Liability) not found"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    created_by = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
    currency = bank_account.currency  # EUR for 101003, MAD for others
    created = 0
    skipped = 0

    # Pre-load existing bank references for duplicate detection
    existing_refs = set(
        JournalEntry.objects.filter(source="bank")
        .exclude(reference="")
        .values_list("reference", flat=True)
    )

    with transaction.atomic():
        for txn in txns:
            debit = Decimal(str(txn.get("debit", 0)))
            credit = Decimal(str(txn.get("credit", 0)))
            if debit == 0 and credit == 0:
                continue

            desc = txn.get("description", "Bank transaction")
            ref = txn.get("reference", "")

            # Skip duplicates by reference
            if ref and ref in existing_refs:
                skipped += 1
                continue

            entry = JournalEntry.objects.create(
                entry_number=_next_entry_number(),
                date=txn["date"],
                description=desc,
                reference=ref,
                source="bank",
                is_posted=True,
                created_by=created_by,
            )

            # Régime de la marge: ALL bank transactions go through
            # Client Funds Liability (240000). No revenue or expense
            # is recognised at the bank import stage.
            if credit > 0:
                # Money IN → DR Bank, CR Client Funds Liability
                JournalEntryLine.objects.create(
                    journal_entry=entry, account=bank_account,
                    debit=credit, credit=Decimal("0"),
                    currency=currency, description=desc,
                )
                JournalEntryLine.objects.create(
                    journal_entry=entry, account=client_funds_acct,
                    debit=Decimal("0"), credit=credit,
                    currency=currency, description=desc,
                )
            else:
                # Money OUT → DR Client Funds Liability, CR Bank
                JournalEntryLine.objects.create(
                    journal_entry=entry, account=client_funds_acct,
                    debit=debit, credit=Decimal("0"),
                    currency=currency, description=desc,
                )
                JournalEntryLine.objects.create(
                    journal_entry=entry, account=bank_account,
                    debit=Decimal("0"), credit=debit,
                    currency=currency, description=desc,
                )

            created += 1

        # Save learned description→account mappings for future auto-map
        for txn in mapped_txns:
            desc = txn.get("description", "")
            acct_id = txn.get("account_id")
            if not desc or not acct_id:
                continue
            desc_key = _extract_description_key(desc)
            if not desc_key:
                continue
            acct = accounts_map.get(acct_id)
            if not acct:
                continue
            # Skip system accounts (not useful as learned mappings)
            if acct.code in ("240000", "210000", "420000"):
                continue
            BankDescriptionMapping.objects.update_or_create(
                description_key=desc_key,
                defaults={
                    "account": acct,
                    "example_description": desc[:500],
                },
            )

    # Save upload history record
    dates = [t["date"] for t in all_txns if t.get("date")]
    try:
        BankStatementUpload.objects.create(
            filename=data.get("filename", "bank_statement"),
            bank_account=bank_account,
            date_from=min(dates) if dates else None,
            date_to=max(dates) if dates else None,
            transaction_count=len(all_txns),
            uploaded_by=f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
        )
    except Exception:
        pass  # Don't fail the whole request if history save fails

    return Response({
        "success": True,
        "message": f"{created} journal entries created, {skipped} duplicates skipped",
        "data": {"count": created, "skipped": skipped},
    }, status=status.HTTP_201_CREATED)


# ─── Reports ──────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_trial_balance(request):
    """Trial balance: all accounts with debit/credit totals from posted entries."""
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")

    accounts = Account.objects.filter(is_active=True).select_related("account_type").order_by("code")
    result = []

    for account in accounts:
        filters = {"journal_entry__is_posted": True}
        if date_from:
            filters["journal_entry__date__gte"] = date_from
        if date_to:
            filters["journal_entry__date__lte"] = date_to

        totals = JournalEntryLine.objects.filter(
            account=account, **filters
        ).aggregate(
            total_debit=Sum("debit"),
            total_credit=Sum("credit"),
        )

        debit = totals["total_debit"] or Decimal("0")
        credit = totals["total_credit"] or Decimal("0")

        if debit == 0 and credit == 0:
            continue

        balance = debit - credit
        result.append({
            "account_id": str(account.id),
            "account_code": account.code,
            "account_name": account.name,
            "account_type": account.account_type.name,
            "debit": str(debit),
            "credit": str(credit),
            "balance": str(balance),
        })

    total_dr = sum(Decimal(r["debit"]) for r in result)
    total_cr = sum(Decimal(r["credit"]) for r in result)

    return Response({
        "success": True,
        "data": {
            "rows": result,
            "total_debit": str(total_dr),
            "total_credit": str(total_cr),
            "is_balanced": total_dr == total_cr,
        },
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_profit_and_loss(request):
    """P&L under margin scheme (régime de la marge).

    Revenue = only account 420000 (Revenue - Margin)
    Expenses = only operating expenses (exclude COS hierarchy descendants)
    Pass-through summary from account 240000 (Client Funds Liability)
    TVA = margin × 20/120
    """
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")

    filters = {"journal_entry__is_posted": True}
    if date_from:
        filters["journal_entry__date__gte"] = date_from
    if date_to:
        filters["journal_entry__date__lte"] = date_to

    # Build COS descendant set (to exclude from operating expenses)
    cos_codes = set()
    try:
        cos_root = Account.objects.get(code="510000")
        def _collect_codes(parent):
            cos_codes.add(parent.code)
            for child in Account.objects.filter(parent=parent):
                _collect_codes(child)
        _collect_codes(cos_root)
    except Account.DoesNotExist:
        pass

    # Pass-through summary from Client Funds Liability (240000)
    try:
        client_funds_acct = Account.objects.get(code="240000")
        cf_totals = JournalEntryLine.objects.filter(
            account=client_funds_acct, **filters
        ).aggregate(total_debit=Sum("debit"), total_credit=Sum("credit"))
        cf_debit = cf_totals["total_debit"] or Decimal("0")
        cf_credit = cf_totals["total_credit"] or Decimal("0")
        # Credits = client funds received, Debits = supplier costs + margin recognition
        client_funds_received = cf_credit
        # Supplier costs = debit entries from bank source (not margin_recognition)
        supplier_filters = {**filters, "journal_entry__source": "bank"}
        supplier_totals = JournalEntryLine.objects.filter(
            account=client_funds_acct, **supplier_filters
        ).aggregate(total_debit=Sum("debit"))
        supplier_costs_paid = supplier_totals["total_debit"] or Decimal("0")
        pass_through_margin = client_funds_received - supplier_costs_paid
    except Account.DoesNotExist:
        client_funds_received = Decimal("0")
        supplier_costs_paid = Decimal("0")
        pass_through_margin = Decimal("0")

    # Revenue = only 420000 (Revenue - Margin)
    revenue_rows = []
    revenue_total = Decimal("0")
    try:
        margin_acct = Account.objects.get(code="420000")
        totals = JournalEntryLine.objects.filter(
            account=margin_acct, **filters
        ).aggregate(total_debit=Sum("debit"), total_credit=Sum("credit"))
        dr = totals["total_debit"] or Decimal("0")
        cr = totals["total_credit"] or Decimal("0")
        amount = cr - dr  # Revenue normal balance = credit
        if amount != 0:
            revenue_total = amount
            revenue_rows.append({
                "account_code": margin_acct.code,
                "account_name": margin_acct.name,
                "amount": str(amount),
            })
    except Account.DoesNotExist:
        pass

    # Operating expenses = all Expense accounts EXCEPT COS descendants
    expense_type = AccountType.objects.filter(name="Expense").first()
    operating_rows = []
    operating_total = Decimal("0")
    if expense_type:
        accounts = Account.objects.filter(
            account_type=expense_type, is_active=True
        ).exclude(code__in=cos_codes).order_by("code")
        for acc in accounts:
            totals = JournalEntryLine.objects.filter(
                account=acc, **filters
            ).aggregate(total_debit=Sum("debit"), total_credit=Sum("credit"))
            dr = totals["total_debit"] or Decimal("0")
            cr = totals["total_credit"] or Decimal("0")
            if dr == 0 and cr == 0:
                continue
            amount = dr - cr  # Expense normal balance = debit
            operating_total += amount
            operating_rows.append({
                "account_code": acc.code,
                "account_name": acc.name,
                "amount": str(amount),
            })

    net_income = revenue_total - operating_total

    # TVA calculation (margin scheme)
    # TVA collectée = margin × 20/120 (VAT is inside the TTC margin)
    # TVA déductible on eligible operating expenses × 20/120
    TVA_ELIGIBLE_CODES = {"592302", "592305", "592306", "592308"}

    tva_eligible_total = Decimal("0")
    tva_eligible_rows = []
    for r in operating_rows:
        if r["account_code"] in TVA_ELIGIBLE_CODES:
            tva_eligible_total += Decimal(r["amount"])
            tva_eligible_rows.append(r)

    # TVA collectée uses GROSS margin (TTC), not net revenue (which has TVA already deducted)
    tva_collectee = (pass_through_margin * Decimal("20") / Decimal("120")) if pass_through_margin > 0 else Decimal("0")
    tva_deductible_ops = (tva_eligible_total * Decimal("20") / Decimal("120")) if tva_eligible_total > 0 else Decimal("0")
    tva_a_payer = tva_collectee - tva_deductible_ops

    return Response({
        "success": True,
        "data": {
            "pass_through": {
                "client_funds_received": str(client_funds_received),
                "supplier_costs_paid": str(supplier_costs_paid),
                "margin": str(pass_through_margin),
            },
            "revenue": {"rows": revenue_rows, "total": str(revenue_total)},
            "operating_expenses": {"rows": operating_rows, "total": str(operating_total)},
            "net_income": str(net_income),
            "tva": {
                "formula": "margin × 20/120",
                "margin_ttc": str(pass_through_margin),
                "collectee": str(tva_collectee),
                "deductible_ops": str(tva_deductible_ops),
                "deductible_ops_rows": tva_eligible_rows,
                "a_payer": str(tva_a_payer),
            },
        },
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_balance_sheet(request):
    """Balance Sheet: Assets = Liabilities + Equity."""
    date_to = request.query_params.get("date_to")

    filters = {"journal_entry__is_posted": True}
    if date_to:
        filters["journal_entry__date__lte"] = date_to

    def get_section(type_name):
        account_type = AccountType.objects.filter(name=type_name).first()
        if not account_type:
            return [], Decimal("0")
        accounts = Account.objects.filter(
            account_type=account_type, is_active=True
        ).order_by("code")
        rows = []
        total = Decimal("0")
        for acc in accounts:
            totals = JournalEntryLine.objects.filter(
                account=acc, **filters
            ).aggregate(total_debit=Sum("debit"), total_credit=Sum("credit"))
            dr = totals["total_debit"] or Decimal("0")
            cr = totals["total_credit"] or Decimal("0")
            if dr == 0 and cr == 0:
                continue
            balance = dr - cr if account_type.normal_balance == "debit" else cr - dr
            total += balance
            rows.append({
                "account_code": acc.code,
                "account_name": acc.name,
                "balance": str(balance),
            })
        return rows, total

    asset_rows, asset_total = get_section("Asset")
    liability_rows, liability_total = get_section("Liability")
    equity_rows, equity_total = get_section("Equity")

    return Response({
        "success": True,
        "data": {
            "assets": {"rows": asset_rows, "total": str(asset_total)},
            "liabilities": {"rows": liability_rows, "total": str(liability_total)},
            "equity": {"rows": equity_rows, "total": str(equity_total)},
            "liabilities_plus_equity": str(liability_total + equity_total),
            "is_balanced": asset_total == (liability_total + equity_total),
        },
    })


# ─── Margin Recognition & TVA ────────────────────────────────────────────────

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def margin_recognition_preview(request):
    """Preview margin recognition calculation for a date range."""
    date_from = request.data.get("date_from")
    date_to = request.data.get("date_to")

    if not date_from or not date_to:
        return Response(
            {"success": False, "message": "date_from and date_to are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        client_funds_acct = Account.objects.get(code="240000")
    except Account.DoesNotExist:
        return Response(
            {"success": False, "message": "Account 240000 (Client Funds Liability) not found"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Calculate from posted bank-source entries on account 240000
    filters = {
        "journal_entry__is_posted": True,
        "journal_entry__source": "bank",
        "journal_entry__date__gte": date_from,
        "journal_entry__date__lte": date_to,
    }
    totals = JournalEntryLine.objects.filter(
        account=client_funds_acct, **filters
    ).aggregate(total_debit=Sum("debit"), total_credit=Sum("credit"))

    client_funds_received = totals["total_credit"] or Decimal("0")
    supplier_costs_paid = totals["total_debit"] or Decimal("0")
    margin = client_funds_received - supplier_costs_paid
    tva = (margin * Decimal("20") / Decimal("120")) if margin > 0 else Decimal("0")

    # Check if entries already exist for this period
    existing_mr = JournalEntry.objects.filter(
        source="margin_recognition",
        date__gte=date_from,
        date__lte=date_to,
    ).count()
    existing_tva = JournalEntry.objects.filter(
        source="tva_margin",
        date__gte=date_from,
        date__lte=date_to,
    ).count()

    already_exists = existing_mr > 0 or existing_tva > 0

    return Response({
        "success": True,
        "data": {
            "client_funds_received": str(client_funds_received),
            "supplier_costs_paid": str(supplier_costs_paid),
            "margin": str(margin),
            "tva": str(tva),
            "already_exists": already_exists,
            "existing_margin_entries": existing_mr,
            "existing_tva_entries": existing_tva,
        },
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def margin_recognition_create(request):
    """Create margin recognition and TVA journal entries."""
    date_from = request.data.get("date_from")
    date_to = request.data.get("date_to")
    entry_date = request.data.get("entry_date")

    if not date_from or not date_to or not entry_date:
        return Response(
            {"success": False, "message": "date_from, date_to, and entry_date are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        client_funds_acct = Account.objects.get(code="240000")
        revenue_acct = Account.objects.get(code="420000")
        tva_acct = Account.objects.get(code="210000")
    except Account.DoesNotExist as e:
        return Response(
            {"success": False, "message": f"Required account not found: {e}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Calculate margin
    filters = {
        "journal_entry__is_posted": True,
        "journal_entry__source": "bank",
        "journal_entry__date__gte": date_from,
        "journal_entry__date__lte": date_to,
    }
    totals = JournalEntryLine.objects.filter(
        account=client_funds_acct, **filters
    ).aggregate(total_debit=Sum("debit"), total_credit=Sum("credit"))

    client_funds_received = totals["total_credit"] or Decimal("0")
    supplier_costs_paid = totals["total_debit"] or Decimal("0")
    margin = client_funds_received - supplier_costs_paid

    if margin <= 0:
        return Response(
            {"success": False, "message": f"Margin is {margin}. No entries to create."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    tva = margin * Decimal("20") / Decimal("120")
    created_by = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username

    with transaction.atomic():
        # 1. Margin Recognition: DR 240000, CR 420000
        mr_entry = JournalEntry.objects.create(
            entry_number=_next_entry_number(),
            date=entry_date,
            description=f"Margin recognition {date_from} to {date_to}",
            reference=f"MR:{date_from}:{date_to}",
            source="margin_recognition",
            source_id=f"{date_from}:{date_to}",
            is_posted=True,
            created_by=created_by,
        )
        JournalEntryLine.objects.create(
            journal_entry=mr_entry,
            account=client_funds_acct,
            debit=margin,
            credit=Decimal("0"),
            description=f"Margin recognition {date_from} to {date_to}",
        )
        JournalEntryLine.objects.create(
            journal_entry=mr_entry,
            account=revenue_acct,
            debit=Decimal("0"),
            credit=margin,
            description=f"Margin recognition {date_from} to {date_to}",
        )

        # 2. TVA on Margin: DR 420000, CR 210000
        tva_entry = JournalEntry.objects.create(
            entry_number=_next_entry_number(),
            date=entry_date,
            description=f"TVA on margin {date_from} to {date_to} ({margin} x 20/120)",
            reference=f"TVA:{date_from}:{date_to}",
            source="tva_margin",
            source_id=f"{date_from}:{date_to}",
            is_posted=True,
            created_by=created_by,
        )
        JournalEntryLine.objects.create(
            journal_entry=tva_entry,
            account=revenue_acct,
            debit=tva,
            credit=Decimal("0"),
            description=f"TVA on margin {date_from} to {date_to}",
        )
        JournalEntryLine.objects.create(
            journal_entry=tva_entry,
            account=tva_acct,
            debit=Decimal("0"),
            credit=tva,
            description=f"TVA on margin {date_from} to {date_to}",
        )

    return Response({
        "success": True,
        "message": "Margin recognition and TVA entries created",
        "data": {
            "margin_entry": _entry_dict(mr_entry),
            "tva_entry": _entry_dict(tva_entry),
        },
    }, status=status.HTTP_201_CREATED)


# ─── CMR Sync ────────────────────────────────────────────────────────────────

# Bank name → account code (mirrors sync_cmr.py)
_CMR_BANK_MAP = {
    "CIH300": "101001",
    "CIH600": "101002",
    "CIH600-DH": "101002",
    "BP09": "101004",
}

_CMR_CATEGORY_ACCOUNT_MAP = {
    "accommodation": "523300",
    "hotel": "523300",
    "restaurant": "533100",
    "transport": "553800",
    "transportation": "553800",
    "guide": "533600",
    "entry": "514000",
    "entries": "514000",
    "package": "510000",
}

_CMR_FALLBACK_ACCOUNT_CODE = "510000"
_CMR_ACCOUNTS_RECEIVABLE_CODE = "120000"
_CMR_CLIENT_FUNDS_CODE = "240000"


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def cmr_sync(request):
    """Sync invoices and supplier payments from CMR into journal entries."""
    since_str = request.data.get("since")
    if not since_str:
        since_str = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()

    try:
        from .cmr_client import CMRClient
        client = CMRClient()
        invoices = client.get_invoices(since=since_str)
        payments = client.get_supplier_payments(since=since_str)
    except Exception as e:
        return Response(
            {"success": False, "message": f"Failed to fetch CMR events: {e}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Pre-load caches
    account_cache = {a.code: a for a in Account.objects.filter(is_active=True)}
    mapping_cache = {
        m.cmr_supplier_id: m.account
        for m in SupplierAccountMapping.objects.select_related("account")
    }

    created_by = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
    created_invoices = 0
    created_payments = 0
    skipped = 0
    details = []

    ar_acct = account_cache.get(_CMR_ACCOUNTS_RECEIVABLE_CODE)
    cf_acct = account_cache.get(_CMR_CLIENT_FUNDS_CODE)

    with transaction.atomic():
        # Process invoices
        for inv in invoices:
            inv_id = inv.get("id", "")
            source_id = inv_id

            if JournalEntry.objects.filter(source="cmr_invoice", source_id=source_id).exists():
                skipped += 1
                details.append(f"SKIP invoice {inv.get('invoice_number')} — already exists")
                continue

            total = Decimal(str(inv.get("total", "0")))
            if total <= 0:
                skipped += 1
                continue

            if not ar_acct or not cf_acct:
                details.append(f"SKIP invoice {inv.get('invoice_number')} — missing AR/CF accounts")
                skipped += 1
                continue

            inv_date = inv.get("invoice_date") or inv.get("updated_at", "")[:10]
            description = (
                f"CMR Invoice {inv.get('invoice_number', '')} — "
                f"{inv.get('customer_name', '')} — {inv.get('opportunity_title', '')}"
            )

            entry = JournalEntry.objects.create(
                entry_number=_next_entry_number(),
                date=inv_date,
                description=description,
                reference=f"CMR:{inv.get('invoice_number', '')}",
                source="cmr_invoice",
                source_id=source_id,
                is_posted=True,
                created_by=created_by,
            )
            # DR line: Accounts Receivable for full total
            JournalEntryLine.objects.create(
                journal_entry=entry, account=ar_acct,
                debit=total, credit=Decimal("0"), description=description,
            )
            # CR lines: split by category from line_items
            line_items = inv.get("line_items") or []
            category_totals = {}
            for li in line_items:
                cat = (li.get("category") or "").strip()
                amt = Decimal(str(li.get("amount", 0)))
                if amt > 0 and cat:
                    category_totals[cat] = category_totals.get(cat, Decimal("0")) + amt

            if category_totals:
                inv_num = inv.get("invoice_number", "")
                cr_sum = Decimal("0")
                cat_items = list(category_totals.items())
                for i, (cat, cat_amt) in enumerate(cat_items):
                    # Last category line: adjust to ensure balance
                    if i == len(cat_items) - 1:
                        cat_amt = total - cr_sum
                    cr_sum += cat_amt
                    cr_desc = f"CMR Invoice {inv_num} — {cat}"
                    JournalEntryLine.objects.create(
                        journal_entry=entry, account=cf_acct,
                        debit=Decimal("0"), credit=cat_amt, description=cr_desc,
                    )
            else:
                # Fallback: single CR line (no categories on line items)
                JournalEntryLine.objects.create(
                    journal_entry=entry, account=cf_acct,
                    debit=Decimal("0"), credit=total, description=description,
                )
            created_invoices += 1
            details.append(f"Created invoice JE: {inv.get('invoice_number')} = {total}")

            # Payment JE if invoice is paid
            if inv.get("status") == "paid" and inv.get("payment_date"):
                payment_source_id = f"{inv_id}:payment"
                if JournalEntry.objects.filter(source="cmr_invoice", source_id=payment_source_id).exists():
                    continue

                bank_name = inv.get("bank_account", "")
                bank_code = _CMR_BANK_MAP.get(bank_name)
                bank_acct = account_cache.get(bank_code) if bank_code else None
                if not bank_acct:
                    bank_acct = account_cache.get("101001")
                if not bank_acct:
                    continue

                pay_desc = f"Payment received — Invoice {inv.get('invoice_number', '')} — {inv.get('customer_name', '')}"
                entry = JournalEntry.objects.create(
                    entry_number=_next_entry_number(),
                    date=inv.get("payment_date"),
                    description=pay_desc,
                    reference=f"CMR:{inv.get('invoice_number', '')}:payment",
                    source="cmr_invoice",
                    source_id=payment_source_id,
                    is_posted=True,
                    created_by=created_by,
                )
                JournalEntryLine.objects.create(
                    journal_entry=entry, account=bank_acct,
                    debit=total, credit=Decimal("0"), description=pay_desc,
                )
                JournalEntryLine.objects.create(
                    journal_entry=entry, account=ar_acct,
                    debit=Decimal("0"), credit=total, description=pay_desc,
                )

        # Process supplier payments
        for pmt in payments:
            opp_id = pmt.get("opportunity_id", "")
            category = pmt.get("category", "")
            index = pmt.get("index", 0)
            source_id = f"{opp_id}:{category}:{index}"

            if JournalEntry.objects.filter(source="cmr_payment", source_id=source_id).exists():
                skipped += 1
                continue

            try:
                amount = Decimal(str(pmt.get("amount", 0)))
            except Exception:
                skipped += 1
                continue

            if amount <= 0:
                skipped += 1
                continue

            pmt_date = pmt.get("payment_date") or datetime.now(timezone.utc).strftime("%Y-%m-%d")
            supplier_name = pmt.get("supplier_name", "Unknown")

            description = (
                f"Supplier payment — {supplier_name} — "
                f"{pmt.get('notes', '')} — Opp: {pmt.get('opportunity_title', '')}"
            ).strip(" —")

            cat_lower = category.lower().strip()
            parent_code = _CMR_CATEGORY_ACCOUNT_MAP.get(cat_lower, _CMR_FALLBACK_ACCOUNT_CODE)
            expense_acct = account_cache.get(parent_code)
            if not expense_acct:
                expense_acct = account_cache.get(_CMR_FALLBACK_ACCOUNT_CODE)
            if not expense_acct:
                skipped += 1
                continue

            bank_name = pmt.get("bank_account", "")
            bank_code = _CMR_BANK_MAP.get(bank_name)
            bank_acct = account_cache.get(bank_code) if bank_code else None
            if not bank_acct:
                bank_acct = account_cache.get("101001")
            if not bank_acct:
                skipped += 1
                continue

            entry = JournalEntry.objects.create(
                entry_number=_next_entry_number(),
                date=pmt_date,
                description=description,
                reference=f"CMR:{opp_id}:{category}:{index}",
                source="cmr_payment",
                source_id=source_id,
                is_posted=True,
                created_by=created_by,
            )
            JournalEntryLine.objects.create(
                journal_entry=entry, account=expense_acct,
                debit=amount, credit=Decimal("0"), description=description,
            )
            JournalEntryLine.objects.create(
                journal_entry=entry, account=bank_acct,
                debit=Decimal("0"), credit=amount, description=description,
            )
            created_payments += 1
            details.append(f"Created payment JE: {supplier_name} ({category}) = {amount}")

    return Response({
        "success": True,
        "message": f"Sync complete: {created_invoices} invoices, {created_payments} payments created",
        "data": {
            "invoices_created": created_invoices,
            "payments_created": created_payments,
            "skipped": skipped,
            "details": details,
        },
    })


# ─── CMR Contact Sync ────────────────────────────────────────────────────────

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def cmr_sync_contacts(request):
    """Sync clients and suppliers from CMR into Customer and Supplier tables."""
    from sales.models import Customer
    from purchases.models import Supplier
    from sales.views import _next_customer_code
    from purchases.views import _next_supplier_code

    try:
        from .cmr_client import CMRClient
        client = CMRClient()
        cmr_clients = client.get_clients()
        cmr_suppliers = client.get_suppliers()
    except Exception as e:
        return Response(
            {"success": False, "message": f"Failed to fetch CMR contacts: {e}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    clients_created = 0
    clients_updated = 0
    suppliers_created = 0
    suppliers_updated = 0
    skipped = 0
    details = []

    with transaction.atomic():
        # ── Sync Clients → Customers ──
        # Build lookup caches
        cmr_id_to_customer = {}
        name_to_customer = {}
        for c in Customer.objects.all():
            if c.cmr_id:
                cmr_id_to_customer[c.cmr_id] = c
            name_to_customer[c.name.lower()] = c

        for cli in cmr_clients:
            cmr_id = cli.get("id", "")
            name = (cli.get("name") or "").strip()
            if not name:
                skipped += 1
                details.append(f"SKIP client — no name (cmr_id={cmr_id})")
                continue

            # Match by cmr_id first, then by name
            customer = cmr_id_to_customer.get(cmr_id)
            matched_by = "cmr_id" if customer else None

            if not customer:
                customer = name_to_customer.get(name.lower())
                matched_by = "name" if customer else None

            if customer:
                # Update existing
                if not customer.cmr_id and cmr_id:
                    customer.cmr_id = cmr_id
                customer.name = name
                if cli.get("email"):
                    customer.email = cli["email"]
                if cli.get("phone"):
                    customer.phone = cli["phone"]
                if cli.get("address"):
                    customer.address = cli["address"]
                if cli.get("tax_id"):
                    customer.tax_id = cli["tax_id"]
                customer.save()
                clients_updated += 1
                details.append(f"Updated customer: {name} (matched by {matched_by})")
            else:
                # Create new
                code = _next_customer_code()
                Customer.objects.create(
                    code=code,
                    name=name,
                    email=cli.get("email", ""),
                    phone=cli.get("phone", ""),
                    address=cli.get("address", ""),
                    tax_id=cli.get("tax_id", ""),
                    cmr_id=cmr_id,
                )
                clients_created += 1
                details.append(f"Created customer: {name} ({code})")

        # ── Sync Suppliers ──
        cmr_id_to_supplier = {}
        name_to_supplier = {}
        for s in Supplier.objects.all():
            if s.cmr_id:
                cmr_id_to_supplier[s.cmr_id] = s
            name_to_supplier[s.name.lower()] = s

        for sup in cmr_suppliers:
            cmr_id = sup.get("id", "")
            name = (sup.get("name") or "").strip()
            if not name:
                skipped += 1
                details.append(f"SKIP supplier — no name (cmr_id={cmr_id})")
                continue

            supplier = cmr_id_to_supplier.get(cmr_id)
            matched_by = "cmr_id" if supplier else None

            if not supplier:
                supplier = name_to_supplier.get(name.lower())
                matched_by = "name" if supplier else None

            if supplier:
                if not supplier.cmr_id and cmr_id:
                    supplier.cmr_id = cmr_id
                supplier.name = name
                if sup.get("email"):
                    supplier.email = sup["email"]
                if sup.get("phone"):
                    supplier.phone = sup["phone"]
                if sup.get("address"):
                    supplier.address = sup["address"]
                if sup.get("tax_id"):
                    supplier.tax_id = sup["tax_id"]
                supplier.save()
                suppliers_updated += 1
                details.append(f"Updated supplier: {name} (matched by {matched_by})")
            else:
                code = _next_supplier_code()
                Supplier.objects.create(
                    code=code,
                    name=name,
                    email=sup.get("email", ""),
                    phone=sup.get("phone", ""),
                    address=sup.get("address", ""),
                    tax_id=sup.get("tax_id", ""),
                    cmr_id=cmr_id,
                )
                suppliers_created += 1
                details.append(f"Created supplier: {name} ({code})")

    return Response({
        "success": True,
        "message": (
            f"Contact sync complete: "
            f"{clients_created} clients created, {clients_updated} updated, "
            f"{suppliers_created} suppliers created, {suppliers_updated} updated"
        ),
        "data": {
            "clients_created": clients_created,
            "clients_updated": clients_updated,
            "suppliers_created": suppliers_created,
            "suppliers_updated": suppliers_updated,
            "skipped": skipped,
            "details": details,
        },
    })


# ─── CMR Client List (proxy for frontend) ────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def cmr_clients_list(request):
    """Proxy CMR clients list to the frontend."""
    try:
        from .cmr_client import CMRClient
        client = CMRClient()
        cmr_clients = client.get_clients(search=request.query_params.get("search"))
    except Exception as e:
        return Response(
            {"success": False, "message": f"Failed to fetch CMR clients: {e}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Mark which ones already exist locally
    from sales.models import Customer
    existing_cmr_ids = set(
        Customer.objects.exclude(cmr_id="").values_list("cmr_id", flat=True)
    )

    results = []
    for cli in cmr_clients:
        # CMR returns company_name/contact_person instead of name
        name = cli.get("company_name") or cli.get("name") or ""
        contact = cli.get("contact_person", "")
        display_name = f"{name} ({contact})" if contact and contact != name else name
        city = cli.get("city", "")
        country = cli.get("country", "")
        address = cli.get("address") or ", ".join(filter(None, [city, country]))
        results.append({
            "id": cli.get("id", ""),
            "name": display_name,
            "email": cli.get("email", ""),
            "phone": cli.get("phone", ""),
            "address": address,
            "tax_id": cli.get("tax_id", ""),
            "currency": cli.get("preferred_currency", "EUR"),
            "already_imported": cli.get("id", "") in existing_cmr_ids,
        })

    return Response({
        "success": True,
        "data": {"clients": results},
    })


# ─── CMR Supplier List (proxy for frontend) ──────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def cmr_suppliers_list(request):
    """Proxy CMR suppliers list to the frontend."""
    try:
        from .cmr_client import CMRClient
        client = CMRClient()
        cmr_suppliers = client.get_suppliers(search=request.query_params.get("search"))
    except Exception as e:
        return Response(
            {"success": False, "message": f"Failed to fetch CMR suppliers: {e}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Mark which ones already exist locally
    from purchases.models import Supplier
    existing_cmr_ids = set(
        Supplier.objects.exclude(cmr_id="").values_list("cmr_id", flat=True)
    )

    results = []
    for sup in cmr_suppliers:
        name = sup.get("company_name") or sup.get("name") or ""
        addr_parts = filter(None, [
            sup.get("address_line1") or "",
            sup.get("address_line2") or "",
            sup.get("postal_code") or "",
            sup.get("city") or "",
            sup.get("country") or "",
        ])
        address = ", ".join(addr_parts)
        results.append({
            "id": sup.get("id", ""),
            "name": name,
            "email": sup.get("email", ""),
            "phone": sup.get("phone", ""),
            "address": address,
            "category": sup.get("category", ""),
            "currency": "MAD",
            "already_imported": sup.get("id", "") in existing_cmr_ids,
        })

    return Response({
        "success": True,
        "data": {"suppliers": results},
    })


# ─── Google Places API ─────────────────────────────────────────────────────────


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def places_autocomplete(request):
    """Google Places Autocomplete (New API) — called directly."""
    query = request.query_params.get("q", "").strip()
    if not query or len(query) < 2:
        return Response({"success": True, "data": {"predictions": []}})

    from django.conf import settings
    import requests as http_requests

    api_key = settings.GOOGLE_PLACES_API_KEY
    if not api_key:
        return Response({"success": False, "message": "GOOGLE_PLACES_API_KEY not configured"}, status=500)

    region = request.query_params.get("region", "").strip()
    url = "https://places.googleapis.com/v1/places:autocomplete"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
    }
    body = {"input": query, "languageCode": "en"}
    if region:
        body["includedRegionCodes"] = [region]

    try:
        resp = http_requests.post(url, json=body, headers=headers, timeout=5)
        data = resp.json()
        predictions = []
        for s in data.get("suggestions", []):
            p = s.get("placePrediction")
            if p:
                predictions.append({
                    "place_id": p.get("placeId"),
                    "description": p.get("text", {}).get("text", ""),
                })
        return Response({"success": True, "data": {"predictions": predictions}})
    except Exception as e:
        return Response({"success": False, "message": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def places_details(request):
    """Google Places Details (New API) — called directly."""
    place_id = request.query_params.get("place_id", "").strip()
    if not place_id:
        return Response({"success": False, "message": "place_id required"}, status=400)

    from django.conf import settings
    import requests as http_requests

    api_key = settings.GOOGLE_PLACES_API_KEY
    if not api_key:
        return Response({"success": False, "message": "GOOGLE_PLACES_API_KEY not configured"}, status=500)

    url = f"https://places.googleapis.com/v1/places/{place_id}"
    headers = {
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": "displayName,formattedAddress,internationalPhoneNumber,nationalPhoneNumber,addressComponents,websiteUri",
    }
    try:
        resp = http_requests.get(url, headers=headers, timeout=5)
        data = resp.json()

        city = ""
        country = ""
        postal_code = ""
        state = ""
        for comp in data.get("addressComponents", []):
            types = comp.get("types", [])
            if "locality" in types:
                city = comp.get("longText", "")
            elif "country" in types:
                country = comp.get("longText", "")
            elif "postal_code" in types:
                postal_code = comp.get("longText", "")
            elif "administrative_area_level_1" in types:
                state = comp.get("longText", "")

        result = {
            "name": data.get("displayName", {}).get("text", ""),
            "address": data.get("formattedAddress", ""),
            "phone": data.get("internationalPhoneNumber", "") or data.get("nationalPhoneNumber", ""),
            "website": data.get("websiteUri", ""),
            "city": city,
            "country": country,
            "postal_code": postal_code,
            "state": state,
        }
        return Response({"success": True, "data": result})
    except Exception as e:
        return Response({"success": False, "message": str(e)}, status=500)


# ─── CMR Invoice Import ───────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def cmr_invoices_list(request):
    """List CMR invoices available for import into the bookkeeping Invoices module."""
    try:
        from .cmr_client import CMRClient
        client = CMRClient()
        cmr_invoices = client.get_invoices(fetch_all=True)
    except Exception as e:
        return Response(
            {"success": False, "message": f"Failed to fetch CMR invoices: {e}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Only show invoices that have been sent to finance
    cmr_invoices = [inv for inv in cmr_invoices if inv.get("sent_to_finance")]

    # Check which invoice numbers are already imported locally
    from sales.models import Invoice as SalesInvoice
    existing_numbers = set(
        SalesInvoice.objects.values_list("invoice_number", flat=True)
    )

    results = []
    for inv in cmr_invoices:
        inv_num = inv.get("invoice_number", "")
        results.append({
            "id": inv.get("id", ""),
            "invoice_number": inv_num,
            "invoice_date": inv.get("invoice_date", ""),
            "due_date": inv.get("due_date", ""),
            "customer_name": inv.get("customer_name", ""),
            "customer_id": inv.get("customer_id", ""),
            "opportunity_title": inv.get("opportunity_title", ""),
            "trip_reference": inv.get("trip_reference", ""),
            "total": inv.get("total", "0"),
            "currency": inv.get("currency", "MAD"),
            "status": inv.get("status", ""),
            "line_items": inv.get("line_items", []),
            "already_imported": inv_num in existing_numbers,
        })

    return Response({
        "success": True,
        "data": {"invoices": results},
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def cmr_invoice_import(request):
    """Import a single CMR invoice as a draft Invoice record."""
    from sales.models import Customer, Invoice as SalesInvoice, InvoiceLine
    from sales.views import _next_customer_code
    from accounts.models import Account

    cmr_invoice_id = request.data.get("cmr_invoice_id")
    if not cmr_invoice_id:
        return Response(
            {"success": False, "message": "cmr_invoice_id is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        from .cmr_client import CMRClient
        client = CMRClient()
        inv = client.get_invoice(cmr_invoice_id)
    except Exception as e:
        return Response(
            {"success": False, "message": f"Failed to fetch CMR invoice: {e}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    if not inv:
        return Response(
            {"success": False, "message": "Invoice not found in CMR"},
            status=status.HTTP_404_NOT_FOUND,
        )

    inv_number = inv.get("invoice_number", "")

    # Check duplicate
    if SalesInvoice.objects.filter(invoice_number=inv_number).exists():
        return Response(
            {"success": False, "message": f"Invoice {inv_number} already exists"},
            status=status.HTTP_409_CONFLICT,
        )

    # Match or create customer
    cmr_customer_id = inv.get("customer_id", "")
    customer_name = inv.get("customer_name", "Unknown")

    customer = None
    if cmr_customer_id:
        customer = Customer.objects.filter(cmr_id=cmr_customer_id).first()
    if not customer:
        customer = Customer.objects.filter(name__iexact=customer_name).first()
    if not customer:
        customer = Customer.objects.create(
            code=_next_customer_code(),
            name=customer_name,
            cmr_id=cmr_customer_id,
        )

    # Resolve account 240000
    try:
        account_240000 = Account.objects.get(code="240000")
    except Account.DoesNotExist:
        return Response(
            {"success": False, "message": "Account 240000 (Client Funds Liability) not found"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    total = Decimal(str(inv.get("total", "0")))
    inv_date = inv.get("invoice_date", "")
    due_date = inv.get("due_date", "") or inv_date
    opp_title = inv.get("opportunity_title", "")
    trip_ref = inv.get("trip_reference", "")
    notes = f"Imported from CMR. Opportunity: {opp_title}. Trip: {trip_ref}"
    currency = inv.get("currency", "MAD")

    with transaction.atomic():
        sales_invoice = SalesInvoice.objects.create(
            invoice_number=inv_number,
            customer=customer,
            date=inv_date,
            due_date=due_date,
            status="draft",
            subtotal=total,
            tax_amount=Decimal("0"),
            total=total,
            currency=currency,
            notes=notes,
            created_by=f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
        )

        line_items = inv.get("line_items") or []
        if line_items:
            for li in line_items:
                amount = Decimal(str(li.get("amount", 0)))
                cat = (li.get("category") or "").strip()
                desc = (li.get("description") or "").strip()
                line_desc = f"[{cat}] {desc}" if cat else desc
                InvoiceLine.objects.create(
                    invoice=sales_invoice,
                    description=line_desc or "CMR line item",
                    quantity=Decimal("1"),
                    unit_price=amount,
                    account=account_240000,
                    tax_code=None,
                )
        else:
            InvoiceLine.objects.create(
                invoice=sales_invoice,
                description=notes,
                quantity=Decimal("1"),
                unit_price=total,
                account=account_240000,
                tax_code=None,
            )

    return Response({
        "success": True,
        "message": f"Invoice {inv_number} imported as draft",
        "data": {"invoice_id": str(sales_invoice.id), "invoice_number": inv_number},
    }, status=status.HTTP_201_CREATED)


# ─── Bank Statement Upload History ───────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def bank_statement_history(request):
    """Return list of previous bank statement uploads."""
    uploads = BankStatementUpload.objects.select_related("bank_account").order_by("-uploaded_at")[:100]
    return Response({
        "success": True,
        "data": {
            "uploads": [
                {
                    "id": str(u.id),
                    "filename": u.filename,
                    "bank_account_code": u.bank_account.code,
                    "bank_account_name": u.bank_account.name,
                    "date_from": u.date_from.isoformat() if u.date_from else None,
                    "date_to": u.date_to.isoformat() if u.date_to else None,
                    "transaction_count": u.transaction_count,
                    "uploaded_by": u.uploaded_by,
                    "uploaded_at": u.uploaded_at.isoformat() if u.uploaded_at else None,
                }
                for u in uploads
            ],
        },
    })


# ─── Journal Entries Export (server-side CSV) ────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def journal_entries_export(request):
    """Export journal entries (with lines) as a flat CSV."""
    qs = JournalEntry.objects.prefetch_related("lines__account").all()

    if request.query_params.get("date_from"):
        qs = qs.filter(date__gte=request.query_params["date_from"])
    if request.query_params.get("date_to"):
        qs = qs.filter(date__lte=request.query_params["date_to"])
    if request.query_params.get("source"):
        qs = qs.filter(source=request.query_params["source"])
    if request.query_params.get("is_posted"):
        qs = qs.filter(is_posted=request.query_params["is_posted"].lower() == "true")
    if request.query_params.get("search"):
        q = request.query_params["search"]
        qs = qs.filter(
            Q(entry_number__icontains=q) | Q(description__icontains=q) | Q(reference__icontains=q)
        )

    qs = qs.order_by("-date", "-entry_number")

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="journal_entries.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow([
        "entry_number", "date", "description", "reference", "source", "is_posted",
        "line_account_code", "line_account_name", "line_debit", "line_credit",
        "line_description", "total_debit", "total_credit",
    ])

    for entry in qs:
        lines = entry.lines.select_related("account").all()
        if lines:
            for line in lines:
                writer.writerow([
                    entry.entry_number, entry.date.isoformat(),
                    entry.description, entry.reference, entry.source,
                    "Yes" if entry.is_posted else "No",
                    line.account.code, line.account.name,
                    str(line.debit), str(line.credit), line.description,
                    str(entry.total_debit), str(entry.total_credit),
                ])
        else:
            writer.writerow([
                entry.entry_number, entry.date.isoformat(),
                entry.description, entry.reference, entry.source,
                "Yes" if entry.is_posted else "No",
                "", "", "", "", "",
                str(entry.total_debit), str(entry.total_credit),
            ])

    return response


# ─── Bank Transactions (flat view) ───────────────────────────────────────────

def _bank_txn_queryset(request):
    """Build filtered queryset for bank transaction lines."""
    qs = JournalEntryLine.objects.filter(
        journal_entry__source="bank",
        account__code__startswith="10",
    ).select_related("journal_entry", "account").order_by(
        "-journal_entry__date", "-journal_entry__entry_number"
    )

    if request.query_params.get("bank_account_id"):
        qs = qs.filter(account_id=request.query_params["bank_account_id"])
    if request.query_params.get("date_from"):
        qs = qs.filter(journal_entry__date__gte=request.query_params["date_from"])
    if request.query_params.get("date_to"):
        qs = qs.filter(journal_entry__date__lte=request.query_params["date_to"])
    if request.query_params.get("search"):
        q = request.query_params["search"]
        qs = qs.filter(
            Q(journal_entry__entry_number__icontains=q) |
            Q(journal_entry__description__icontains=q) |
            Q(journal_entry__reference__icontains=q)
        )

    # Sorting
    sort_field = request.query_params.get("sort", "date")
    sort_dir = request.query_params.get("sort_dir", "desc")
    allowed_sort = {
        "date": "journal_entry__date",
        "entry_number": "journal_entry__entry_number",
        "description": "journal_entry__description",
        "reference": "journal_entry__reference",
        "bank_account": "account__code",
        "debit": "debit",
        "credit": "credit",
    }
    order_field = allowed_sort.get(sort_field, "journal_entry__date")
    prefix = "-" if sort_dir == "desc" else ""
    qs = qs.order_by(f"{prefix}{order_field}", "-journal_entry__date")

    return qs


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def bank_transactions_list(request):
    """Flat list of bank-source transactions (one row per bank-side JE line)."""
    qs = _bank_txn_queryset(request)

    # Summary aggregates on full filtered set
    agg = qs.aggregate(total_debit=Sum("debit"), total_credit=Sum("credit"))
    total_debit = agg["total_debit"] or Decimal("0")
    total_credit = agg["total_credit"] or Decimal("0")

    total_count = qs.count()

    # Pagination
    try:
        page = int(request.query_params.get("page", 1))
    except (ValueError, TypeError):
        page = 1
    try:
        page_size = int(request.query_params.get("page_size", 50))
    except (ValueError, TypeError):
        page_size = 50
    page_size = min(page_size, 200)
    offset = (page - 1) * page_size
    lines = qs[offset:offset + page_size]

    rows = []
    for line in lines:
        je = line.journal_entry
        rows.append({
            "id": str(line.id),
            "entry_number": je.entry_number,
            "date": je.date.isoformat() if hasattr(je.date, "isoformat") else str(je.date),
            "description": je.description,
            "reference": je.reference,
            "debit": str(line.debit),
            "credit": str(line.credit),
            "currency": line.currency,
            "bank_account_id": str(line.account_id),
            "bank_account_code": line.account.code,
            "bank_account_name": line.account.name,
        })

    # Bank accounts for filter dropdown
    bank_accounts = list(
        Account.objects.filter(code__startswith="10", is_active=True)
        .order_by("code")
        .values("id", "code", "name")
    )
    bank_accounts = [
        {"id": str(a["id"]), "code": a["code"], "name": a["name"]}
        for a in bank_accounts
    ]

    return Response({
        "success": True,
        "data": {
            "transactions": rows,
            "count": total_count,
            "total_count": total_count,
            "page": page,
            "page_size": page_size,
            "summary": {
                "total_debit": str(total_debit),
                "total_credit": str(total_credit),
                "net": str(total_debit - total_credit),
            },
            "bank_accounts": bank_accounts,
        },
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def bank_transactions_export(request):
    """Export bank transactions as CSV with BOM."""
    qs = _bank_txn_queryset(request)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="bank_transactions.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow(["Date", "Entry Number", "Description", "Reference", "Bank Account", "Debit", "Credit", "Currency"])

    for line in qs:
        je = line.journal_entry
        writer.writerow([
            je.date.isoformat() if hasattr(je.date, "isoformat") else str(je.date),
            je.entry_number,
            je.description,
            je.reference,
            f"{line.account.code} - {line.account.name}",
            str(line.debit),
            str(line.credit),
            line.currency,
        ])

    return response
