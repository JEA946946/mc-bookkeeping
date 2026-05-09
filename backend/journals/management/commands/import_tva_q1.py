"""Import TVA Q1 2026 bank transactions using margin scheme (régime de la marge).

Patterns:
A. Client payment → DR Bank, CR Client Funds Liability (240000)
B. Supplier payment (COS descendant) → DR Client Funds Liability (240000), CR Bank
C. Operating expense (non-COS) → DR Expense, CR Bank  (unchanged)
D. Margin recognition → DR Client Funds Liability (240000), CR Revenue-Margin (420000)
E. TVA on margin → DR Revenue-Margin (420000), CR VAT Payable (210000)
"""

import uuid
from datetime import date
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from django.db import transaction

from accounts.models import Account, AccountType
from journals.models import JournalEntry, JournalEntryLine


# Bank name → bank account code mapping
BANK_MAP = {
    "BP09": "101004",
    "CIH300": "101001",
    "CIH600": "101002",
    "CIH600-DH": "101002",
}

# Excel account codes that conflict with DB codes.
CODE_REMAP = {
    "100020": "513000",  # La Caisse Entrees → COS under 510000
    "553813": "553819",  # Oncf → use existing DB code 553819
    "553814": "553820",  # Trek → use existing DB code 553820
}

# New accounts to create (code, name, type_id, parent_code)
# type 4 = Revenue, type 5 = Expense
NEW_ACCOUNTS = [
    # Revenue - new clients
    ("410126", "WeTravel", 4, "410000"),
    ("410127", "Hanibal Travel", 4, "410000"),
    ("410128", "Impro Travel", 4, "410000"),
    ("410129", "Air Tours Club", 4, "410000"),
    ("411099", "BERGERREISID", 4, "410000"),
    # Entries - new suppliers
    ("514008", "KTI - Roundtrips", 5, "514000"),
    ("514009", "Chez Naji", 5, "514000"),
    # Accommodation - new hotels (parent: 523300)
    ("523526", "The Au Bout Du Monde", 5, "523300"),
    ("523527", "Riad les oliviers", 5, "523300"),
    ("523528", "Hotel le Tinsouline", 5, "523300"),
    ("523529", "Dar Mouna", 5, "523300"),
    ("523530", "Bivouac Sand Cheggaga", 5, "523300"),
    ("523531", "Hotel les Amandiers", 5, "523300"),
    ("523532", "Oasis Dar Anou", 5, "523300"),
    ("523533", "The Blue Man Camp", 5, "523300"),
    ("523534", "Riad Essaoussan", 5, "523300"),
    ("523535", "Riad Zyo", 5, "523300"),
    ("523536", "Hotel Aruba Boutique", 5, "523300"),
    ("523537", "Riad Bahia Salam", 5, "523300"),
    ("523538", "Riad Mazar Fes", 5, "523300"),
    ("523539", "Riad Milouda", 5, "523300"),
    ("523540", "Hotel Atlantic Agadir", 5, "523300"),
    ("523541", "Riad Monceau", 5, "523300"),
    ("523542", "Riad Diamond", 5, "523300"),
    ("523543", "Riad Mogantique", 5, "523300"),
    ("523544", "Riad Ghali", 5, "523300"),
    ("523545", "Riad Maya", 5, "523300"),
    ("523546", "Riad Ambre et Jasmin", 5, "523300"),
    ("523547", "Agafay Luxury Camp", 5, "523300"),
    ("523548", "La Maison de tanger", 5, "523300"),
    ("523549", "Hotel Oudaya", 5, "523300"),
    ("523550", "Hotel Kenzi Azghor", 5, "523300"),
    ("523551", "Kasbah Merzourga", 5, "523300"),
    ("523552", "Euphoriad", 5, "523300"),
    ("523553", "Camp Lemnouar", 5, "523300"),
    ("523554", "Villa Quieta", 5, "523300"),
    ("523555", "Hotel Albakech", 5, "523300"),
    # Restaurant - new suppliers (parent: 533100)
    ("533189", "Restaurant Le Grand", 5, "533100"),
    ("533190", "Restaurant Narwama", 5, "533100"),
    ("533192", "Restaurant Shakerato", 5, "533100"),
    # Visa (flight/travel related)
    ("533700", "Visa", 5, "510000"),
    # Guide - new guides (parent: 533600)
    ("543626", "Guide Qarouach Hassane", 5, "533600"),
    ("543627", "Guide Hasse Ed Dahni", 5, "533600"),
    ("543628", "Guide Mohammed El Harrak", 5, "533600"),
    ("543629", "Guide Lahcen", 5, "533600"),
    ("543630", "Guide hachimi mohamed", 5, "533600"),
    ("543631", "Guide Tarik Salih", 5, "533600"),
    # Transport - new companies (parent: 553800)
    ("553816", "Zoubir", 5, "553800"),
    ("553817", "Transport Ask For Tour", 5, "553800"),
]

# Salary account reassignments (update existing DB records)
SALARY_FIXES = {
    "572201": "CNSS",
    "572202": "Salary Kaouthar",
}

# Margin scheme account codes
CLIENT_FUNDS_CODE = "240000"
MARGIN_REVENUE_CODE = "420000"
VAT_PAYABLE_CODE = "210000"
COS_ROOT_CODE = "510000"


def safe_float(val):
    """Convert Excel cell value to float, returning 0.0 for non-numeric."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _collect_descendant_codes(root_code):
    """Collect all account codes under a root (including root itself)."""
    codes = set()
    try:
        root = Account.objects.get(code=root_code)
    except Account.DoesNotExist:
        return codes

    def _walk(parent):
        codes.add(parent.code)
        for child in Account.objects.filter(parent=parent):
            _walk(child)

    _walk(root)
    return codes


class Command(BaseCommand):
    help = "Import TVA Q1 2026 bank transactions using margin scheme"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="/app/Files/TVA 2026.xlsx",
            help="Path to TVA Excel file",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without creating entries",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Delete ALL existing journal entries before importing",
        )

    def handle(self, *args, **options):
        filepath = options["file"]
        dry_run = options["dry_run"]
        clear = options["clear"]

        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        self.stdout.write(f"Sheet: {ws.title}, Rows: {ws.max_row}")

        # Step 0: Clear existing entries if requested
        if clear and not dry_run:
            count = JournalEntry.objects.count()
            JournalEntry.objects.all().delete()
            self.stdout.write(self.style.WARNING(
                f"\n--- Cleared {count} existing journal entries ---"
            ))

        # Step 1: Fix salary account names
        self.stdout.write("\n--- Step 1: Fix salary account names ---")
        for code, new_name in SALARY_FIXES.items():
            try:
                acct = Account.objects.get(code=code)
                old_name = acct.name
                if not dry_run:
                    acct.name = new_name
                    acct.save()
                self.stdout.write(f"  Updated {code}: '{old_name}' → '{new_name}'")
            except Account.DoesNotExist:
                self.stdout.write(self.style.WARNING(f"  {code} not found, skipping"))

        # Step 2: Create missing accounts
        self.stdout.write(f"\n--- Step 2: Create {len(NEW_ACCOUNTS)} missing accounts ---")
        created_count = 0
        for code, name, type_id, parent_code in NEW_ACCOUNTS:
            if Account.objects.filter(code=code).exists():
                self.stdout.write(f"  {code} already exists, skipping")
                continue
            parent = None
            if parent_code:
                try:
                    parent = Account.objects.get(code=parent_code)
                except Account.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f"  Parent {parent_code} not found for {code}, creating without parent"
                    ))
            if not dry_run:
                Account.objects.create(
                    code=code,
                    name=name,
                    account_type_id=type_id,
                    parent=parent,
                    currency="MAD",
                )
            created_count += 1
            self.stdout.write(f"  Created {code}: {name}")
        self.stdout.write(f"  Total created: {created_count}")

        # Step 2b: Build COS descendant set for classification
        cos_codes = _collect_descendant_codes(COS_ROOT_CODE)
        self.stdout.write(f"\n  COS hierarchy: {len(cos_codes)} accounts under {COS_ROOT_CODE}")

        # Step 3: Parse Excel rows
        self.stdout.write("\n--- Step 3: Parse Excel transactions ---")
        transactions = []
        skipped = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
            bank_name = row[0]
            op_date = row[1]
            description = row[3]
            reference = row[4]
            acct_code_raw = row[5]
            acct_name = row[6]
            invoice_ref = row[8]
            debit = safe_float(row[9])
            credit = safe_float(row[10])

            # Skip empty rows
            if bank_name is None and acct_code_raw is None:
                continue

            # Skip header-like rows
            if isinstance(acct_code_raw, str) and acct_code_raw == "Code":
                continue

            # Normalize account code
            if acct_code_raw is None:
                errors.append(f"Row {row_idx}: No account code")
                skipped += 1
                continue

            acct_code = str(int(acct_code_raw)) if isinstance(acct_code_raw, float) else str(acct_code_raw)

            # Apply code remapping for conflicts
            db_code = CODE_REMAP.get(acct_code, acct_code)

            # Get bank account code
            bank_code = BANK_MAP.get(str(bank_name)) if bank_name else None
            if not bank_code:
                errors.append(f"Row {row_idx}: Unknown bank '{bank_name}'")
                skipped += 1
                continue

            # Parse date
            if hasattr(op_date, 'date'):
                txn_date = op_date.date()
            elif hasattr(op_date, 'strftime'):
                txn_date = op_date
            else:
                errors.append(f"Row {row_idx}: Invalid date '{op_date}'")
                skipped += 1
                continue

            # Must have either debit or credit
            if debit == 0 and credit == 0:
                errors.append(f"Row {row_idx}: Both debit and credit are 0")
                skipped += 1
                continue

            # Build description
            desc_parts = []
            if description:
                desc_parts.append(str(description))
            if acct_name:
                desc_parts.append(str(acct_name))
            txn_desc = " — ".join(desc_parts) if desc_parts else "Bank transaction"

            ref_str = str(reference) if reference else ""
            inv_ref = str(invoice_ref) if invoice_ref else ""

            # Classify transaction
            # Credit + 41xxxx → client payment (A): DR Bank, CR 240000
            # Debit + COS descendant → supplier payment (B): DR 240000, CR Bank
            # Debit + 100020 (La Caisse Entrees) → supplier payment (B): cash used for COS
            # Debit + 41xxxx → client refund (B): reduces client funds via 240000
            # Debit + other 10xxxx (La Caisse, banks) → asset transfer: DR Asset, CR Bank
            # Debit + non-COS expense → operating expense (C): DR Expense, CR Bank
            is_client_payment = credit > 0 and db_code.startswith("41")
            is_cos = db_code in cos_codes

            if credit > 0:
                txn_type = "client_payment" if is_client_payment else "other_credit"
            elif is_cos:
                txn_type = "supplier_payment"
            elif db_code.startswith("41"):
                txn_type = "supplier_payment"  # client refund reduces client funds
            elif db_code.startswith("10"):
                txn_type = "asset_transfer"  # internal cash/bank transfer
            else:
                txn_type = "operating_expense"

            transactions.append({
                "date": txn_date,
                "description": txn_desc,
                "reference": ref_str,
                "invoice_ref": inv_ref,
                "acct_code": db_code,
                "bank_code": bank_code,
                "debit": Decimal(str(debit)) if debit else Decimal("0"),
                "credit": Decimal(str(credit)) if credit else Decimal("0"),
                "bank_name": str(bank_name),
                "row": row_idx,
                "type": txn_type,
            })

        self.stdout.write(f"  Parsed: {len(transactions)} transactions, {skipped} skipped")
        if errors:
            for e in errors[:10]:
                self.stdout.write(self.style.WARNING(f"  {e}"))
            if len(errors) > 10:
                self.stdout.write(self.style.WARNING(f"  ... and {len(errors)-10} more errors"))

        # Classification summary
        type_counts = {}
        for t in transactions:
            type_counts[t["type"]] = type_counts.get(t["type"], 0) + 1
        self.stdout.write("\n  Classification summary:")
        for k, v in sorted(type_counts.items()):
            self.stdout.write(f"    {k}: {v}")

        # Compute totals per type
        client_funds_received = sum(
            t["credit"] for t in transactions if t["type"] == "client_payment"
        )
        supplier_costs_paid = sum(
            t["debit"] for t in transactions if t["type"] == "supplier_payment"
        )
        opex_total = sum(
            t["debit"] for t in transactions if t["type"] == "operating_expense"
        )
        asset_transfers = sum(
            t["debit"] for t in transactions if t["type"] == "asset_transfer"
        )
        margin = client_funds_received - supplier_costs_paid
        tva_amount = (margin * Decimal("20") / Decimal("120")) if margin > 0 else Decimal("0")

        self.stdout.write(f"\n  Client funds received: {client_funds_received:>14,.2f}")
        self.stdout.write(f"  Supplier costs paid:   {supplier_costs_paid:>14,.2f}")
        self.stdout.write(f"  Margin:                {margin:>14,.2f}")
        self.stdout.write(f"  TVA (margin×20/120):   {tva_amount:>14,.2f}")
        self.stdout.write(f"  Operating expenses:    {opex_total:>14,.2f}")
        self.stdout.write(f"  Asset transfers:       {asset_transfers:>14,.2f}")

        if dry_run:
            self.stdout.write(self.style.SUCCESS("\n--- DRY RUN complete, no entries created ---"))
            return

        # Step 4: Validate all accounts exist
        self.stdout.write("\n--- Step 4: Validate accounts ---")
        all_codes = set()
        for t in transactions:
            all_codes.add(t["acct_code"])
            all_codes.add(t["bank_code"])
        # Also need margin scheme accounts
        all_codes.update([CLIENT_FUNDS_CODE, MARGIN_REVENUE_CODE, VAT_PAYABLE_CODE])

        account_cache = {}
        missing_codes = []
        for code in sorted(all_codes):
            try:
                account_cache[code] = Account.objects.get(code=code)
            except Account.DoesNotExist:
                missing_codes.append(code)

        if missing_codes:
            self.stdout.write(self.style.ERROR(
                f"  MISSING accounts: {', '.join(missing_codes)}"
            ))
            self.stdout.write(self.style.ERROR("  Aborting import. Create these accounts first."))
            return

        self.stdout.write(f"  All {len(all_codes)} accounts validated")

        # Step 5: Create journal entries
        self.stdout.write("\n--- Step 5: Create journal entries (margin scheme) ---")

        # Get next entry number
        last = JournalEntry.objects.order_by("-entry_number").first()
        if last:
            try:
                next_num = int(last.entry_number.replace("JE-", "")) + 1
            except ValueError:
                next_num = 1
        else:
            next_num = 1

        client_funds_acct = account_cache[CLIENT_FUNDS_CODE]
        margin_revenue_acct = account_cache[MARGIN_REVENUE_CODE]
        vat_payable_acct = account_cache[VAT_PAYABLE_CODE]

        created = 0
        with transaction.atomic():
            for txn in transactions:
                entry_number = f"JE-{next_num:06d}"

                # Build reference
                ref_parts = [txn["bank_name"]]
                if txn["reference"]:
                    ref_parts.append(txn["reference"])
                if txn["invoice_ref"]:
                    ref_parts.append(f"Inv:{txn['invoice_ref']}")
                full_ref = " | ".join(ref_parts)

                bank_acct = account_cache[txn["bank_code"]]

                if txn["type"] == "client_payment":
                    # Pattern A: DR Bank, CR Client Funds Liability
                    amount = txn["credit"]
                    entry = JournalEntry.objects.create(
                        entry_number=entry_number,
                        date=txn["date"],
                        description=txn["description"],
                        reference=full_ref,
                        source="bank",
                        is_posted=True,
                        created_by="TVA Q1 Import (margin)",
                    )
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account=bank_acct,
                        debit=amount,
                        credit=Decimal("0"),
                        description=txn["description"],
                    )
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account=client_funds_acct,
                        debit=Decimal("0"),
                        credit=amount,
                        description=txn["description"],
                    )

                elif txn["type"] == "supplier_payment":
                    # Pattern B: DR Client Funds Liability, CR Bank
                    amount = txn["debit"]
                    entry = JournalEntry.objects.create(
                        entry_number=entry_number,
                        date=txn["date"],
                        description=txn["description"],
                        reference=full_ref,
                        source="bank",
                        is_posted=True,
                        created_by="TVA Q1 Import (margin)",
                    )
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account=client_funds_acct,
                        debit=amount,
                        credit=Decimal("0"),
                        description=txn["description"],
                    )
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account=bank_acct,
                        debit=Decimal("0"),
                        credit=amount,
                        description=txn["description"],
                    )

                elif txn["type"] == "operating_expense":
                    # Pattern C: DR Expense, CR Bank (unchanged)
                    amount = txn["debit"]
                    acct = account_cache[txn["acct_code"]]
                    entry = JournalEntry.objects.create(
                        entry_number=entry_number,
                        date=txn["date"],
                        description=txn["description"],
                        reference=full_ref,
                        source="bank",
                        is_posted=True,
                        created_by="TVA Q1 Import (margin)",
                    )
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account=acct,
                        debit=amount,
                        credit=Decimal("0"),
                        description=txn["description"],
                    )
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account=bank_acct,
                        debit=Decimal("0"),
                        credit=amount,
                        description=txn["description"],
                    )

                elif txn["type"] == "asset_transfer":
                    # Internal transfer: DR Asset (cash), CR Bank
                    amount = txn["debit"]
                    acct = account_cache[txn["acct_code"]]
                    entry = JournalEntry.objects.create(
                        entry_number=entry_number,
                        date=txn["date"],
                        description=txn["description"],
                        reference=full_ref,
                        source="bank",
                        is_posted=True,
                        created_by="TVA Q1 Import (margin)",
                    )
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account=acct,
                        debit=amount,
                        credit=Decimal("0"),
                        description=txn["description"],
                    )
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account=bank_acct,
                        debit=Decimal("0"),
                        credit=amount,
                        description=txn["description"],
                    )

                else:
                    # other_credit: non-client credit (rare) — treat same as pattern A
                    # but credit goes to the original account
                    amount = txn["credit"]
                    acct = account_cache[txn["acct_code"]]
                    entry = JournalEntry.objects.create(
                        entry_number=entry_number,
                        date=txn["date"],
                        description=txn["description"],
                        reference=full_ref,
                        source="bank",
                        is_posted=True,
                        created_by="TVA Q1 Import (margin)",
                    )
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account=bank_acct,
                        debit=amount,
                        credit=Decimal("0"),
                        description=txn["description"],
                    )
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account=acct,
                        debit=Decimal("0"),
                        credit=amount,
                        description=txn["description"],
                    )

                created += 1
                next_num += 1

                if created % 100 == 0:
                    self.stdout.write(f"  Created {created} entries...")

            # Pattern D: Margin recognition (if margin > 0)
            if margin > 0:
                entry_number = f"JE-{next_num:06d}"
                entry = JournalEntry.objects.create(
                    entry_number=entry_number,
                    date=date(2026, 3, 31),
                    description="Margin recognition Q1 2026 (régime de la marge)",
                    reference="Margin Q1",
                    source="margin_recognition",
                    is_posted=True,
                    created_by="TVA Q1 Import (margin)",
                )
                JournalEntryLine.objects.create(
                    journal_entry=entry,
                    account=client_funds_acct,
                    debit=margin,
                    credit=Decimal("0"),
                    description="Margin recognized from client funds",
                )
                JournalEntryLine.objects.create(
                    journal_entry=entry,
                    account=margin_revenue_acct,
                    debit=Decimal("0"),
                    credit=margin,
                    description="Margin recognized from client funds",
                )
                created += 1
                next_num += 1
                self.stdout.write(f"  Created margin recognition entry: {margin:,.2f} MAD")

            # Pattern E: TVA on margin (if tva > 0)
            if tva_amount > 0:
                # Round to 2 decimal places
                tva_amount = tva_amount.quantize(Decimal("0.01"))
                entry_number = f"JE-{next_num:06d}"
                entry = JournalEntry.objects.create(
                    entry_number=entry_number,
                    date=date(2026, 3, 31),
                    description="TVA on margin Q1 2026 (margin × 20/120)",
                    reference="TVA Margin Q1",
                    source="tva_margin",
                    is_posted=True,
                    created_by="TVA Q1 Import (margin)",
                )
                JournalEntryLine.objects.create(
                    journal_entry=entry,
                    account=margin_revenue_acct,
                    debit=tva_amount,
                    credit=Decimal("0"),
                    description="TVA collectée on margin",
                )
                JournalEntryLine.objects.create(
                    journal_entry=entry,
                    account=vat_payable_acct,
                    debit=Decimal("0"),
                    credit=tva_amount,
                    description="TVA collectée on margin",
                )
                created += 1
                next_num += 1
                self.stdout.write(f"  Created TVA on margin entry: {tva_amount:,.2f} MAD")

        self.stdout.write(self.style.SUCCESS(
            f"\n  Successfully imported {created} journal entries"
        ))

        # Final summary
        self.stdout.write("\n--- Final Summary ---")
        self.stdout.write(f"  Client funds received: {client_funds_received:>14,.2f} MAD")
        self.stdout.write(f"  Supplier costs paid:   {supplier_costs_paid:>14,.2f} MAD")
        self.stdout.write(f"  Margin:                {margin:>14,.2f} MAD")
        self.stdout.write(f"  TVA (margin×20/120):   {tva_amount:>14,.2f} MAD")
        self.stdout.write(f"  Operating expenses:    {opex_total:>14,.2f} MAD")
        self.stdout.write(f"  Net income:            {(margin - tva_amount - opex_total):>14,.2f} MAD")
